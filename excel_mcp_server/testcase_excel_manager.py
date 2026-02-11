#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel测试用例管理工具
用于读取、写入、修改Metersphere格式的测试用例xlsx文件
支持完整的Excel操作，包括格式保持和中文内容处理
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict, Any, Optional
import copy


class TestCaseExcelManager:
    """测试用例Excel管理器"""
    
    # 标准列定义
    COLUMNS = [
        "用例名称",
        "前置条件",
        "所属模块",
        "步骤描述",
        "预期结果",
        "备注",
        "用例等级",
        "编辑模式"
    ]
    
    def __init__(self, file_path: str):
        """
        初始化管理器
        
        Args:
            file_path: xlsx文件路径
        """
        self.file_path = file_path
        self.wb = None
        self.ws = None
        self.data = None
    
    def open(self):
        """打开现有Excel文件"""
        try:
            self.wb = load_workbook(self.file_path, data_only=False)
            self.ws = self.wb.active
            # 使用pandas读取数据
            self.data = pd.read_excel(self.file_path, sheet_name=0, engine='openpyxl')
            print(f"成功打开文件: {self.file_path}")
            print(f"工作表: {self.ws.title}")
            print(f"数据行数: {len(self.data)}")
            return True
        except Exception as e:
            print(f"打开文件失败: {e}")
            return False
    
    def read_testcases(self) -> List[Dict[str, Any]]:
        """
        读取所有测试用例
        
        Returns:
            测试用例列表，每个用例为字典格式
        """
        if self.data is None:
            self.open()
        
        testcases = []
        for idx, row in self.data.iterrows():
            testcase = {}
            for col in self.COLUMNS:
                value = row.get(col, "")
                # 处理NaN值
                if pd.isna(value):
                    value = ""
                testcase[col] = value
            testcases.append(testcase)
        
        return testcases
    
    def read_testcase_by_name(self, name: str) -> Optional[Dict[str, Any]]:
        """
        根据名称查找测试用例
        
        Args:
            name: 用例名称
        
        Returns:
            找到的测试用例，未找到返回None
        """
        testcases = self.read_testcases()
        for tc in testcases:
            if tc["用例名称"] == name:
                return tc
        return None
    
    def get_cell_value(self, row: int, column: str) -> Any:
        """
        获取指定单元格的值
        
        Args:
            row: 行号（从1开始）
            column: 列号（A, B, C...）
        
        Returns:
            单元格的值
        """
        if self.ws is None:
            return None
        cell = self.ws[f"{column}{row}"]
        return cell.value
    
    def update_cell(self, row: int, column: str, value: Any, preserve_format: bool = True):
        """
        更新指定单元格的值
        
        Args:
            row: 行号（从1开始）
            column: 列号（A, B, C...）
            value: 新值
            preserve_format: 是否保持原有格式
        """
        if self.ws is None:
            raise Exception("工作表未打开")
        
        cell = self.ws[f"{column}{row}"]
        
        if preserve_format:
            # 保持原有格式，只修改值
            old_cell = copy.copy(cell)
            cell.value = value
        else:
            cell.value = value
        
        print(f"更新单元格 {column}{row}: {value}")
    
    def update_testcase(self, row: int, testcase: Dict[str, Any]):
        """
        更新整行测试用例数据
        
        Args:
            row: 行号（从2开始，第1行是表头）
            testcase: 测试用例数据字典
        """
        if self.ws is None:
            raise Exception("工作表未打开")
        
        col_map = {
            "用例名称": "A",
            "前置条件": "B",
            "所属模块": "C",
            "步骤描述": "D",
            "预期结果": "E",
            "备注": "F",
            "用例等级": "G",
            "编辑模式": "H"
        }
        
        for col_name, col_letter in col_map.items():
            if col_name in testcase:
                self.update_cell(row, col_letter, testcase[col_name])
    
    def add_testcase(self, testcase: Dict[str, Any]):
        """
        添加新的测试用例到文件末尾
        
        Args:
            testcase: 测试用例数据字典
        """
        if self.ws is None:
            raise Exception("工作表未打开")
        
        new_row = self.ws.max_row + 1
        
        # 确保所有列都有值
        for col in self.COLUMNS:
            if col not in testcase:
                testcase[col] = ""
        
        self.update_testcase(new_row, testcase)
        print(f"添加新用例到第{new_row}行")
    
    def delete_testcase(self, row: int):
        """
        删除指定行的测试用例
        
        Args:
            row: 行号（从2开始）
        """
        if self.ws is None:
            raise Exception("工作表未打开")
        
        # 删除该行
        self.ws.delete_rows(row, 1)
        print(f"删除第{row}行的测试用例")
    
    def save(self, new_path: Optional[str] = None):
        """
        保存Excel文件
        
        Args:
            new_path: 新文件路径，如果为None则覆盖原文件
        """
        if self.wb is None:
            raise Exception("工作簿未打开")
        
        save_path = new_path if new_path else self.file_path
        self.wb.save(save_path)
        print(f"文件已保存: {save_path}")
    
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
            print("工作簿已关闭")
    
    def __enter__(self):
        """上下文管理器入口"""
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()
        return False


class TestCaseExcelCreator:
    """测试用例Excel创建器"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def _get_header_style(self):
        """获取表头样式"""
        # 字体：宋体，14号，加粗
        font = Font(name='宋体', size=14, bold=True)
        
        # 对齐：水平居中，垂直居中，自动换行
        alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # 边框：细线
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        return {'font': font, 'alignment': alignment, 'border': thin_border}
    
    def _apply_header_style(self):
        """应用表头样式到第一行"""
        header_style = self._get_header_style()
        
        for col_idx, col_name in enumerate(TestCaseExcelManager.COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.font = header_style['font']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
    
    def create_new_file(self, testcases: List[Dict[str, Any]], output_path: str) -> bool:
        """
        创建新的测试用例Excel文件
        
        Args:
            testcases: 测试用例列表
            output_path: 输出文件路径
        
        Returns:
            是否成功
        """
        try:
            # 创建新工作簿
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "模版"
            
            # 设置表头
            for col_idx, col_name in enumerate(TestCaseExcelManager.COLUMNS, 1):
                self.ws.cell(row=1, column=col_idx, value=col_name)
            
            # 应用表头样式
            self._apply_header_style()
            
            # 写入数据
            for row_idx, testcase in enumerate(testcases, start=2):
                for col_idx, col_name in enumerate(TestCaseExcelManager.COLUMNS, 1):
                    value = testcase.get(col_name, "")
                    # 处理NaN值
                    if pd.isna(value):
                        value = ""
                    self.ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 设置列宽（根据内容自动调整）
            self._adjust_column_width()
            
            # 保存文件
            self.wb.save(output_path)
            print(f"成功创建文件: {output_path}")
            print(f"包含 {len(testcases)} 条测试用例")
            return True
            
        except Exception as e:
            print(f"创建文件失败: {e}")
            return False
    
    def _adjust_column_width(self):
        """根据内容自动调整列宽"""
        for col_idx in range(1, len(TestCaseExcelManager.COLUMNS) + 1):
            # 查找该列中最长的内容
            max_length = 0
            for row in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col_idx)
                if cell.value:
                    value = str(cell.value)
                    # 考虑换行，按行计算
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, max_line_length)
            
            # 设置列宽（每个中文字符占2个单位）
            column_letter = chr(64 + col_idx)  # A, B, C...
            self.ws.column_dimensions[column_letter].width = max_length * 1.2 + 4


def validate_testcase(testcase: Dict[str, Any]) -> List[str]:
    """
    验证测试用例数据的完整性
    
    Args:
        testcase: 测试用例数据
    
    Returns:
        错误信息列表，空列表表示验证通过
    """
    errors = []
    
    # 必填字段检查
    required_fields = ["用例名称", "所属模块", "步骤描述", "预期结果", "用例等级"]
    for field in required_fields:
        if field not in testcase or not testcase[field]:
            errors.append(f"缺少必填字段: {field}")
    
    # 用例等级检查
    if testcase.get("用例等级") not in ["P0", "P1", "P2"]:
        errors.append(f"用例等级无效: {testcase.get('用例等级')}")
    
    # 编辑模式检查
    if testcase.get("编辑模式") not in ["TEXT", "RICH_TEXT", "其他"]:
        errors.append(f"编辑模式无效: {testcase.get('编辑模式')}")
    
    return errors


def print_testcase(testcase: Dict[str, Any], idx: Optional[int] = None):
    """
    打印测试用例信息
    
    Args:
        testcase: 测试用例数据
        idx: 序号（可选）
    """
    if idx is not None:
        print(f"\n{'='*80}")
        print(f"测试用例 #{idx}")
        print(f"{'='*80}")
    
    for col in TestCaseExcelManager.COLUMNS:
        value = testcase.get(col, "")
        if value:
            print(f"{col}:")
            if col in ["步骤描述", "预期结果"]:
                # 多行内容显示
                lines = str(value).split('\n')
                for line in lines:
                    print(f"  {line}")
            else:
                print(f"  {value}")


if __name__ == "__main__":
    # 示例使用
    
    # 1. 读取现有文件
    print("=" * 80)
    print("示例1: 读取现有Excel文件")
    print("=" * 80)
    
    manager = TestCaseExcelManager(r"d:/Code/autotest/用例生成/示例模板/Metersphere_case_IT资产盘点.xlsx")
    
    with manager:
        # 读取所有测试用例
        testcases = manager.read_testcases()
        print(f"\n共读取到 {len(testcases)} 条测试用例")
        
        # 打印前3条用例
        for idx, tc in enumerate(testcases[:3], 1):
            print_testcase(tc, idx)
        
        # 查找特定用例
        print(f"\n查找用例: 【冒烟】PC端资产盘点完整流程")
        found_tc = manager.read_testcase_by_name("【冒烟】PC端资产盘点完整流程")
        if found_tc:
            print("找到用例:")
            print_testcase(found_tc)
    
    # 2. 创建新文件示例
    print("\n" + "=" * 80)
    print("示例2: 创建新的Excel文件")
    print("=" * 80)
    
    new_testcases = [
        {
            "用例名称": "【冒烟】示例测试用例1",
            "前置条件": "用户已登录系统",
            "所属模块": "/示例模块/子模块/功能",
            "步骤描述": "[1]执行步骤1\n[2]执行步骤2\n[3]执行步骤3",
            "预期结果": "[1]预期结果1\n[2]预期结果2\n[3]预期结果3",
            "备注": "",
            "用例等级": "P0",
            "编辑模式": "TEXT"
        },
        {
            "用例名称": "示例测试用例2",
            "前置条件": "存在测试数据",
            "所属模块": "/示例模块/子模块",
            "步骤描述": "[1]执行测试操作",
            "预期结果": "[1]验证测试结果",
            "备注": "这是一个P1级用例",
            "用例等级": "P1",
            "编辑模式": "TEXT"
        }
    ]
    
    creator = TestCaseExcelCreator()
    creator.create_new_file(
        new_testcases,
        r"d:/Code/autotest/用例生成/示例输出/new_testcases.xlsx"
    )